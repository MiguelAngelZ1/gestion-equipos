#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
from datetime import datetime
from pathlib import Path
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configurar salida para soportar caracteres especiales en consolas Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# ================================
# CONFIGURACIÓN DE RUTAS INTELIGENTE
# ================================
IS_WINDOWS = os.name == 'nt'

def get_cloud_paths():
    if not IS_WINDOWS: return []
    
    home = Path(os.path.expanduser("~"))
    rutas_encontradas = []
    
    print("[DEBUG] Escaneando perfil de usuario para detectar carpetas de nube...")
    
    try:
        # 1. Buscar OneDrive (dinámico: "OneDrive", "OneDrive - Personal", etc.)
        for folder in home.iterdir():
            if folder.is_dir() and "onedrive" in folder.name.lower():
                ruta_od = folder / "Exportaciones"
                rutas_encontradas.append(ruta_od)
                print(f"[DEBUG] Detectado OneDrive en: {folder.name}")
        
        # 2. Buscar Google Drive (Unidad G: o Carpeta en Home)
        drive_g = Path("G:/Mi unidad/Exportaciones")
        if drive_g.parent.exists():
            rutas_encontradas.append(drive_g)
            print("[DEBUG] Detectada Unidad G: de Google Drive")
        
        for folder in home.iterdir():
            if folder.is_dir() and "google drive" in folder.name.lower():
                # Buscar subcarpetas "Mi unidad" o "My Drive"
                found_sub = False
                try:
                    for sub in folder.iterdir():
                        if sub.is_dir() and sub.name.lower() in ["mi unidad", "my drive"]:
                            rutas_encontradas.append(sub / "Exportaciones")
                            found_sub = True
                            print(f"[DEBUG] Detectado Google Drive (Local) en: {sub.name}")
                            break
                except: pass
                if not found_sub:
                    rutas_encontradas.append(folder / "Exportaciones")
                    print(f"[DEBUG] Detectado Google Drive (Local) en: {folder.name}")

        # 3. Fallbacks Estáticos (si no se encontró nada dinámicamente)
        fallbacks = [
            Path(r"C:\Users\Miguel Angel Imperio\OneDrive\Exportaciones"),
            Path(r"C:\Users\Miguel Angel Imperio\Mi unidad\Exportaciones")
        ]
        for f in fallbacks:
            if f.parent.exists() and f not in rutas_encontradas:
                rutas_encontradas.append(f)

    except Exception as e:
        print(f"[ERROR] Error durante el escaneo de rutas: {e}")

    # Eliminar duplicados manteniendo el orden
    resultado = list(dict.fromkeys(rutas_encontradas))
    return resultado

RUTAS_OBJETIVO = get_cloud_paths()

def formatear_especificaciones(specs):
    if not specs: return ""
    return "\n".join(f"{s.get('clave', '')}: {s.get('valor', '')}" for s in specs)

def crear_excel(equipos_data, output_path):
    print(f"[INFO] Generando Excel con {len(equipos_data)} registros...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['INE', 'NNE', 'Serie', 'Tipo', 'Estado', 'Responsable', 'Ubicacion', 'Especificaciones']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, eq in enumerate(equipos_data, start=2):
            ws.cell(i, 1, eq.get('ine', '-'))
            ws.cell(i, 2, eq.get('nne', '-'))
            ws.cell(i, 3, eq.get('serie', '-'))
            ws.cell(i, 4, eq.get('tipo', '-'))
            ws.cell(i, 5, eq.get('estado', '-'))
            ws.cell(i, 6, eq.get('responsable', '-'))
            ws.cell(i, 7, eq.get('ubicacion', '-'))

            specs = eq.get('especificaciones', [])
            cell_specs = ws.cell(i, 8, formatear_especificaciones(specs))
            cell_specs.alignment = Alignment(wrap_text=True, vertical="top")
            cell_specs.border = border

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['H'].width = 50
        wb.save(output_path)
        print(f"[OK] Excel guardado en: {output_path}")
        return True
    except Exception as e:
        print(f"[ERROR] Error creando Excel: {e}")
        return False

def copiar_a_ruta_con_nombre(fuente, carpeta_destino, nombre_nuevo):
    try:
        # Asegurar que la carpeta de destino existe
        if not carpeta_destino.exists():
            print(f"[INFO] Creando carpeta de destino: {carpeta_destino}")
            carpeta_destino.mkdir(parents=True, exist_ok=True)
            
        destino_final = carpeta_destino / nombre_nuevo
        shutil.copy2(fuente, destino_final)
        print(f"[OK] Sincronizado en: {destino_final}")
        return True
    except Exception as e:
        print(f"[WARN] No se pudo guardar en {carpeta_destino}: {e}")
        return False

def main():
    print("\n" + "="*50)
    print("INICIANDO EXPORTACION A RUTAS DE NUBE")
    print("="*50)

    if len(sys.argv) < 3:
        print("[ERROR] Faltan argumentos.")
        sys.exit(1)

    temp_origen = Path(sys.argv[1])
    data_source = sys.argv[2]

    # Cargar datos
    try:
        with open(data_source, 'r', encoding='utf-8') as f:
            equipos_data = json.load(f)
    except Exception as e:
        print(f"[ERROR] Fallo al leer datos: {e}")
        sys.exit(1)

    # 1. Crear el nombre de archivo final (ej: Reporte_Equipos_2025-01-24.xlsx)
    fecha_str = datetime.now().strftime("%Y-%m-%d")
    nombre_final = f"Reporte_Equipos_{fecha_str}.xlsx"

    # 2. Crear el archivo Excel en la ruta temporal
    if crear_excel(equipos_data, str(temp_origen)):
        # 3. Si estamos en Windows, realizar las copias a las carpetas de nube
        if IS_WINDOWS:
            exitos = 0
            if not RUTAS_OBJETIVO:
                print("[WARN] No se detectaron carpetas de nube automáticamente.")
            
            for ruta_base in RUTAS_OBJETIVO:
                # Aseguramos que el archivo vaya SIEMPRE dentro de la subcarpeta 'Exportaciones'
                if copiar_a_ruta_con_nombre(temp_origen, ruta_base, nombre_final):
                    exitos += 1
            
            if exitos == 0:
                print("[ERROR] No se pudo guardar en NINGUNA de las carpetas de nube.")
                # Fallback al Escritorio si todo falla
                escritorio = Path(os.path.expanduser("~/Desktop")) / "Respaldo_Exportaciones"
                copiar_a_ruta_con_nombre(temp_origen, escritorio, nombre_final)
        else:
            print("[INFO] Entorno Linux/Nube detectado. Saltando copias locales.")
    
    print("="*50 + "\n")
    print("PROCESO FINALIZADO")
    print("="*50)

if __name__ == "__main__":
    main()
